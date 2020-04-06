#! /usr/bin/python3 
import openpyxl
 #Create a file called config.py and add acesss token in thew same dir, it will be ignored by git by .gitignore
import config
import datetime
from twilio.rest import Client
from openpyxl import Workbook
from datetime import datetime, timedelta, date

wb_fileName = 'Forum_DB.xlsx'
#In the First
def main():
	wb = Workbook()
	sheet = wb.active
	wb = openpyxl.load_workbook(wb_fileName)
	get_date(wb)
	print('From Config: ', config.account_sid, ' ', config.auth_token)

#Send Notifications
#def send_notification(due_date, noOfDays, phoneNum):
#	if noOfDays >= 0:
#		msg = "Kindly note that your subcription expires on " + str(due_date.date()) +", Kindly renew." + "\nPhone: " + phoneNum
#	else:
#	print(msg)

#Send Notifications
def send_notification(due_date, noOfDays, phoneNum):
	countryCode = '+91'
	client = Client(config.account_sid, config.auth_token)
	if noOfDays >= 0:
		msg = "Kindly note that your subcription expires on " + str(due_date) +", Kindly renew." + "\nPhone: " + phoneNum # + "\nPhone: " + phoneNum + " Trillo Test"
	else:
		msg = "Kindly note that your subcription has already expired  " + str(-1 *(due_date)) +" ago, Kindly renew."# + "\nPhone: " + phoneNum + " Trillo Test"

	smsMessageResponse = client.messages.create(
                     body = msg,
                     from_ = config.twilloPhnNum,
                     to = countryCode + phoneNum
                 )
	print(smsMessageResponse.sid)

def get_date(wb):
	StartRow, EndRow, DueDateCol, PhoneNumCol, MembershipCol = 2, 8, 3, 4, 6
	due_date = 0 
	current = wb.active

	for i in range(StartRow, EndRow, 1):
		#payment_date  
		payment_date = datetime.strptime(str(current.cell(row = i, column = DueDateCol).value), '%Y-%m-%d %H:%M:%S')
		today = datetime.today()
		ndays = (today - payment_date).days
		#print('Due Date: ', due_date)
		#print('days: ', ndays)
		phoneNum = str(current.cell(row = i, column = PhoneNumCol).value)
		Membership_type = str(current.cell(row = i, column = MembershipCol).value)
		if(Membership_type == 'Annual'):
			if(ndays == 357 or ndays >= 365):
				due_date = ndays - 365
				print('DueDate:', str(due_date), str(phoneNum))
				send_notification(due_date, ndays, phoneNum)
		elif(Membership_type == 'BiAnnual'):
			if(ndays == 723 or ndays >= 730):
				due_date = ndays - 730
				print('DueDate:', str(due_date), str(phoneNum))
				send_notification(due_date, ndays, phoneNum)
		elif(Membership_type == 'Five Years'):
			if(ndays == 1818 or ndays >= 1825):
				due_date = ndays - 1825
				print('DueDate:', str(due_date), str(phoneNum))
				send_notification(due_date, ndays, phoneNum)
		else:
			print('Invalid Membership Type')

			
if __name__=="__main__":
	main()
